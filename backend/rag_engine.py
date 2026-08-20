"""
RAG engine for report Q&A: turns real xlsx/xls/csv/pdf reports into text
chunks, retrieves the most relevant ones for a question with a local
TF-IDF cosine-similarity search (no external embedding API required —
Cohere is used automatically instead for retrieval if COHERE_API_KEY is
configured, since it gives materially better results), and generates the
final answer with Cohere, Gemini, or Groq (your choice, via AI_PROVIDER),
strictly grounded in the retrieved chunks plus recent conversation history.

No fabricated / hardcoded report content lives in this file. If nothing
relevant is found, the model is instructed to say so plainly.
"""
import csv
import json
import logging
import math
import re
from collections import Counter
from pathlib import Path

try:
    import numpy as np
except ImportError:
    np = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import cohere
except ImportError:
    cohere = None

try:
    import requests
except ImportError:
    requests = None

log = logging.getLogger("rag_engine")

COHERE_CHAT_MODEL_DEFAULT = "command-r-plus"
GEMINI_MODEL_DEFAULT = "gemini-2.5-flash"
GEMINI_URL_TMPL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
GROQ_MODEL_DEFAULT = "openai/gpt-oss-120b"
GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"

# How many past question/answer pairs to include as context for follow-up
# questions (e.g. "rephrase my last question", "what about the other one").
MAX_HISTORY_TURNS = 6

_TOKEN_RE = re.compile(r"[a-z0-9]+")

_SYSTEM_PROMPT_TMPL = (
    "You are the TAXILLA AI Assistant, embedded in the TAXILLA Compliance & Reconciliation Platform. "
    "Answer the user's question using ONLY the TAXILLA SYSTEM REPORT DATA below — never invent numbers, "
    "names, or facts that are not present in it. If the answer isn't in the data, state clearly "
    "that the data is not available in current TAXILLA system reports.\n\n"
    "CRITICAL RULES FOR RESPONSES:\n"
    "1. Never mention 'emails', 'mailboxes', 'gmail', 'outlook', 'attachments', or 'uploads'.\n"
    "2. NEVER mention specific file names (e.g. '.xlsx', '.csv', '.pdf', 'Scope 2...'), file paths, or data origin sources. Present all data directly as native system reconciliation figures.\n"
    "3. If multiple reports exist in the data, automatically perform side-by-side comparisons, "
    "reconcile variances, and highlight key differences across reports.\n"
    "4. Format your text answer professionally using concise bullet points, executive summaries, and markdown tables.\n"
    "5. Use precomputed statistics directly whenever present.\n"
    "6. VISUALIZATION MANDATE: Whenever your answer contains numbers, totals, emissions, quantities, costs, or metric values, YOU MUST ALWAYS GENERATE A VALID VISUALIZATION CHART! "
    "When comparing multiple approaches or categories (e.g. Location Based vs Market Based), USE A STACKED BAR CHART by setting stacked=true, seriesKeys=['Location Based', 'Market Based'], and chartData containing metrics for each category.\n\n"
    "{history_section}"
    "TAXILLA SYSTEM REPORT DATA:\n{context}\n\n"
    "QUESTION: {query}\n\n"
    "Respond with ONLY raw JSON, no markdown fences, no preamble, matching exactly this shape:\n"
    '{{"answer": "string", "chartType": "bar" | "pie" | null, "stacked": boolean or null, "seriesKeys": ["string"] or null, "chartTitle": "string or null", '
    '"chartData": [{{"name": "string", "value": number}}] or null}}'
)

REPORT_ENHANCEMENT_PROMPT_TMPL = (
    "You are the TAXILLA AI Assistant generating an executive HTML dashboard report for the "
    "TAXILLA Compliance & Reconciliation Platform. Transform the RAW REPORT DATA below into a "
    "polished, dynamic, dashboard-style HTML report — not a plain table dump.\n\n"

    "DATA GROUNDING RULES (do not violate):\n"
    "1. Use ONLY numbers, labels, and facts present in RAW REPORT DATA. Never invent, estimate, "
    "or round in a way that changes a figure.\n"
    "2. Never mention 'email', 'mailbox', 'gmail', 'outlook', 'attachment', 'upload', file names, "
    "or file extensions anywhere in the output. Present everything as native TAXILLA system data.\n"
    "3. If a narrative field already contains an embedded <img> tag (base64 data-URI), preserve "
    "that tag exactly as-is, in its original section — do not alter, recompress, or drop it.\n"
    "4. Never expose how the report was assembled — no 'Sheet: X' labels, no worksheet/tab names, "
    "no mention of rows, columns, cells, or parsing. Use clean, business-appropriate section titles "
    "derived from the content itself (e.g. a worksheet about facilities becomes a 'Facility Breakdown' "
    "section, not 'Sheet: Facility Data').\n"
    "5. Omit rows/records that carry no real information — every value blank, zero, '0%', or '$-'. "
    "These are unused template placeholders, not data points, and must not appear in tables, KPI "
    "cards, or the narrative as if they were real figures.\n\n"

    "STRUCTURE TO GENERATE:\n"
    "A. Header banner — report title, period/date range if present in data, generated timestamp.\n"
    "B. KPI summary strip — 3 to 6 cards, each with: a small inline SVG icon (not an external image "
    "URL), the metric label, the value exactly as in the data, and a trend/delta indicator ONLY if "
    "a comparison figure exists in the data (never fabricate a trend).\n"
    "C. Executive summary — 2 to 4 short paragraphs of plain-language insight synthesized from the "
    "data (biggest movers, notable ratios, anything that stands out). Every claim must trace back "
    "to a number in RAW REPORT DATA.\n"
    "D. Chart section — pick 1 to 3 chart types (bar/line/pie/doughnut) that best fit the data shape "
    "you were given. Use Chart.js loaded via CDN. Populate datasets ONLY from values present in the data.\n"
    "E. Detailed tables — every row/column from the raw data must still appear somewhere, styled "
    "(zebra striping, sticky header, right-aligned numerics).\n"
    "F. Footer — 'Confidential Executive Dashboard Report Generated by TAXILLA AI Assistant'.\n\n"

    "If RAW REPORT DATA is empty or has no real rows, respond with a short, styled HTML notice "
    "saying no report data is available yet — do not invent a report.\n\n"

    "VISUAL REQUIREMENTS:\n"
    "- Single self-contained HTML file: inline <style> only, one "
    "<script src='https://cdn.jsdelivr.net/npm/chart.js'></script> tag for charts, no other "
    "external assets.\n"
    "- Modern CSS: grid/flexbox layout, a consistent palette, generous whitespace, readable type "
    "scale, mobile-responsive, print-friendly.\n"
    "- Do not use Lorem Ipsum or placeholder charts — every element must map to real data or be omitted.\n\n"

    "ADDITIONAL INSTRUCTIONS FROM USER:\n{instructions}\n\n"

    "RAW REPORT DATA:\n{report_context}\n\n"

    "Respond with ONLY the complete HTML document, starting with <!DOCTYPE html> and ending with "
    "</html>. No markdown fences, no commentary, no explanation before or after."
)


def _strip_html_fences(text):
    """Some providers wrap HTML in ```html ... ``` fences despite instructions
    not to. Strip those, and trim to the <!DOCTYPE ...>...</html> span if the
    model added any stray preamble/postamble text around it."""
    if not text:
        return text
    t = text.strip()
    if t.startswith("```"):
        t = re.sub(r"^```(?:html)?\s*", "", t)
        t = re.sub(r"\s*```$", "", t)
    lower = t.lower()
    start = lower.find("<!doctype")
    if start == -1:
        start = lower.find("<html")
    end = lower.rfind("</html>")
    if start != -1 and end != -1:
        t = t[start:end + len("</html>")]
    return t.strip()


def _tokenize(text):
    return _TOKEN_RE.findall((text or "").lower())


# Belt-and-suspenders safety net: the system prompt already instructs the
# model never to mention file names, extensions, or mail-related origin
# words, but LLMs occasionally slip (quoting a filename verbatim from the
# context, or saying "the attached file"). This scrubs any that leak
# through before the answer ever reaches the person, since the prompt
# alone isn't a hard guarantee.
_METADATA_LEAK_PATTERNS = [
    re.compile(r"\b[\w\-. ]+?\.(xlsx|xls|csv|pdf)\b", re.IGNORECASE),
    re.compile(r"\b(email|e-mail|mailbox|gmail|outlook|attachment|attached file|uploaded file)\b", re.IGNORECASE),
]


def _scrub_metadata_leakage(answer_text):
    if not answer_text:
        return answer_text
    cleaned = answer_text
    for pattern in _METADATA_LEAK_PATTERNS:
        cleaned = pattern.sub("the report", cleaned)
    return cleaned


def _scrub_json_answer(raw_json_text):
    """The model's raw response is expected to be a JSON string with an
    "answer" field (see _SYSTEM_PROMPT_TMPL). Parses it, scrubs just that
    field, and re-serializes — leaves the string untouched (rather than
    raising) if it isn't valid JSON, since the caller/frontend already
    has its own tolerant fallback parsing for that case."""
    try:
        text = raw_json_text.strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
        data = json.loads(text)
        if isinstance(data, dict) and "answer" in data:
            data["answer"] = _scrub_metadata_leakage(data["answer"])
            return json.dumps(data)
        return raw_json_text
    except Exception:
        return raw_json_text


def _build_history_section(history):
    """history: list of {"question": str, "answer": str} pairs, oldest
    first. Returns a prompt section (or "" if no history), trimmed to the
    most recent MAX_HISTORY_TURNS turns so the prompt doesn't grow
    unbounded over a long conversation."""
    if not history:
        return ""
    recent = history[-MAX_HISTORY_TURNS:]
    lines = ["CONVERSATION HISTORY (most recent last — use this to understand follow-up "
              "questions like 'rephrase that' or 'what about the other one', but still ground "
              "every factual answer in REPORT DATA below, not in what you said earlier):"]
    for turn in recent:
        q = (turn.get("question") or "").strip()
        a = (turn.get("answer") or "").strip()
        if q:
            lines.append(f"User: {q}")
        if a:
            lines.append(f"Assistant: {a}")
    lines.append("")  # blank line separator before REPORT DATA
    return "\n".join(lines) + "\n\n"


class LocalTfidfIndex:
    """Zero-dependency TF-IDF + cosine similarity retriever. Good enough
    for report-style spreadsheets/PDFs without needing an embeddings API."""

    def __init__(self, chunks):
        self.chunks = chunks
        self.doc_tokens = [_tokenize(c["text"]) for c in chunks]
        df = Counter()
        for toks in self.doc_tokens:
            for t in set(toks):
                df[t] += 1
        n_docs = max(len(chunks), 1)
        self.idf = {t: math.log((n_docs + 1) / (freq + 1)) + 1 for t, freq in df.items()}
        self.doc_vecs = [self._vectorize(toks) for toks in self.doc_tokens]

    def _vectorize(self, tokens):
        tf = Counter(tokens)
        vec = {t: (count / max(len(tokens), 1)) * self.idf.get(t, 0.0) for t, count in tf.items()}
        norm = math.sqrt(sum(v * v for v in vec.values())) or 1.0
        return {t: v / norm for t, v in vec.items()}

    @staticmethod
    def _cosine(a, b):
        if len(a) > len(b):
            a, b = b, a
        return sum(v * b.get(t, 0.0) for t, v in a.items())

    def top_k(self, query, k=6):
        if not self.chunks:
            return []
        q_vec = self._vectorize(_tokenize(query))
        scored = [(self._cosine(q_vec, dv), i) for i, dv in enumerate(self.doc_vecs)]
        scored.sort(key=lambda x: x[0], reverse=True)
        return [self.chunks[i] for score, i in scored[:k] if score > 0] or self.chunks[:k]


class ReportRAGEngine:
    def __init__(self, cohere_api_key: str = None, gemini_api_key: str = None, groq_api_key: str = None,
                 ai_provider: str = "cohere", cohere_model: str = None, gemini_model: str = None,
                 groq_model: str = None):
        self.cohere_api_key = cohere_api_key
        self.gemini_api_key = gemini_api_key
        self.groq_api_key = groq_api_key
        self.ai_provider = (ai_provider or "cohere").lower()
        self.cohere_model = cohere_model or COHERE_CHAT_MODEL_DEFAULT
        self.gemini_model = gemini_model or GEMINI_MODEL_DEFAULT
        self.groq_model = groq_model or GROQ_MODEL_DEFAULT
        self.co_client = None
        if cohere_api_key and cohere:
            try:
                self.co_client = cohere.Client(cohere_api_key)
            except Exception as e:
                log.warning("Could not initialize Cohere client: %s", e)

    # ---------- Precomputed statistics (accurate sums/counts/averages) ----------

    def _compute_sheet_stats(self, path: Path, sheet_name: str, filename: str,
                              max_numeric_cols=6, max_categorical_cols=4, max_pivot_cols=2):
        """Precomputes real totals/counts/averages and a couple of group
        breakdowns with pandas, so aggregation questions don't rely on the
        AI manually adding up rows from raw text — which is unreliable
        even when every row is present in context."""
        if pd is None:
            return None
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        except Exception:
            return None
        if df.empty or len(df.columns) == 0:
            return None
        df.columns = [str(c).strip() for c in df.columns]

        numeric_cols = []
        for c in df.columns:
            coerced = pd.to_numeric(df[c], errors="coerce")
            if coerced.notna().sum() >= max(3, len(df) * 0.5):
                numeric_cols.append(c)

        categorical_cols = []
        for c in df.columns:
            if c in numeric_cols:
                continue
            nunique = df[c].nunique(dropna=True)
            if 2 <= nunique <= 30:
                categorical_cols.append((c, nunique))
        categorical_cols.sort(key=lambda x: x[1])

        priority_keywords = ("score", "quantity", "amount", "total", "emission",
                              "cost", "liability", "consumption", "value")
        prioritized_numeric = sorted(
            numeric_cols,
            key=lambda c: (0 if any(k in c.lower() for k in priority_keywords) else 1),
        )[:max_numeric_cols]

        if not prioritized_numeric and not categorical_cols:
            return None

        lines = [f"PRECOMPUTED STATISTICS for sheet '{sheet_name}' in {filename}",
                 "(Use these numbers directly for totals/sums/counts/averages instead of "
                 "manually adding up individual rows.)"]

        if prioritized_numeric:
            lines.append("\nColumn totals:")
            for c in prioritized_numeric:
                s = pd.to_numeric(df[c], errors="coerce")
                lines.append(
                    f"- {c}: Sum={s.sum():.2f}, Count={int(s.notna().sum())}, "
                    f"Average={s.mean():.2f}, Min={s.min():.2f}, Max={s.max():.2f}"
                )

        top_categorical = categorical_cols[:max_categorical_cols]
        for c, _ in top_categorical:
            counts = df[c].value_counts(dropna=True)
            lines.append(f"\nRow counts by '{c}':")
            for val, cnt in counts.items():
                lines.append(f"- {val}: {cnt}")

        if len(top_categorical) >= 2 and prioritized_numeric:
            key_a, key_b = top_categorical[0][0], top_categorical[1][0]
            for num_col in prioritized_numeric[:max_pivot_cols]:
                try:
                    s = pd.to_numeric(df[num_col], errors="coerce")
                    pivot = df.assign(_val=s).groupby([key_a, key_b])["_val"].sum()
                    lines.append(f"\n'{num_col}' totals by '{key_a}' and '{key_b}':")
                    for (a, b), v in pivot.items():
                        if v:
                            lines.append(f"- {a} / {b}: {v:.2f}")
                except Exception:
                    continue

        return "\n".join(lines)

    # ---------- Loading & chunking real report content ----------

    def load_xlsx_chunks(self, path: Path, filename: str, rows_per_chunk: int = 12):
        if not openpyxl or not path.exists():
            return []
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            log.error("Error loading workbook %s: %s", path, e)
            return []

        chunks = []
        for sheet in wb.worksheets:
            for merged_range in list(sheet.merged_cells.ranges):
                top_left_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                sheet.unmerge_cells(str(merged_range))
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        sheet.cell(row=r, column=c).value = top_left_value

            stats_text = self._compute_sheet_stats(path, sheet.title, filename)
            if stats_text:
                chunks.append({"source": filename, "sheet": sheet.title, "text": stats_text})

            row_lines = []
            for row in sheet.iter_rows():
                cells = [c.value for c in row if c.value is not None and str(c.value).strip() != ""]
                if cells:
                    row_lines.append(", ".join(str(v) for v in cells))

            for i in range(0, len(row_lines), rows_per_chunk):
                group = row_lines[i:i + rows_per_chunk]
                if not group:
                    continue
                text = f"Report: {filename}\nSheet: {sheet.title}\n" + "\n".join(group)
                chunks.append({"source": filename, "sheet": sheet.title, "text": text})
        return chunks

    def load_csv_chunks(self, path: Path, filename: str, rows_per_chunk: int = 20):
        if not path.exists():
            return []
        try:
            with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
                rows = [", ".join(str(cell) for cell in row if str(cell).strip() != "") for row in csv.reader(f)]
                rows = [r for r in rows if r]
        except Exception as e:
            log.error("Error loading CSV %s: %s", path, e)
            return []

        chunks = []
        for i in range(0, len(rows), rows_per_chunk):
            group = rows[i:i + rows_per_chunk]
            if not group:
                continue
            text = f"Report: {filename}\n" + "\n".join(group)
            chunks.append({"source": filename, "sheet": None, "text": text})
        return chunks

    def load_pdf_text_chunks(self, extracted_text: str, filename: str, chars_per_chunk: int = 1600):
        if not extracted_text or not extracted_text.strip():
            return []
        chunks = []
        for i in range(0, len(extracted_text), chars_per_chunk):
            piece = extracted_text[i:i + chars_per_chunk].strip()
            if piece:
                chunks.append({"source": filename, "sheet": None, "text": f"Report: {filename}\n{piece}"})
        return chunks

    # ---------- Retrieval ----------

    def _embed_cohere(self, texts, input_type):
        response = self.co_client.embed(texts=texts, model="embed-english-v3.0", input_type=input_type)
        return np.array(response.embeddings)

    def retrieve(self, query: str, chunks, top_k: int = 6):
        if not chunks:
            return []
        if self.co_client and np is not None:
            try:
                doc_embeddings = self._embed_cohere([c["text"] for c in chunks], "search_document")
                query_embedding = self._embed_cohere([query], "search_query")[0]
                norms = np.linalg.norm(doc_embeddings, axis=1) * np.linalg.norm(query_embedding)
                norms[norms == 0] = 1e-10
                scores = (doc_embeddings @ query_embedding) / norms
                top_indices = np.argsort(scores)[::-1][:top_k]
                return [chunks[i] for i in top_indices]
            except Exception as e:
                log.warning("Cohere retrieval failed, falling back to local TF-IDF: %s", e)
        return LocalTfidfIndex(chunks).top_k(query, k=top_k)

    # ---------- Generation — Cohere, Gemini, or Groq, grounded strictly in retrieved chunks ----------

    def _generate_cohere(self, prompt: str):
        if not self.co_client:
            return None
        try:
            # Low, fixed temperature — the earlier default (unset, so each
            # provider's own default applied, typically ~0.7-1.0) meant the
            # exact same question could get noticeably different answers
            # from one ask to the next. This is a factual-reporting
            # assistant, not a creative one, so consistency matters more
            # than variety here.
            response = self.co_client.chat(message=prompt, model=self.cohere_model, temperature=0.1)
            return (response.text or "").strip()
        except Exception as e:
            log.error("Cohere chat call failed: %s", e)
            return None

    def _generate_gemini(self, prompt: str, timeout: int = 30):
        if not self.gemini_api_key or requests is None:
            return None
        url = GEMINI_URL_TMPL.format(model=self.gemini_model)
        try:
            resp = requests.post(
                url,
                params={"key": self.gemini_api_key},
                headers={"content-type": "application/json"},
                json={"contents": [{"parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.1}},
                timeout=timeout,
            )
            resp.raise_for_status()
            data = resp.json()
            candidates = data.get("candidates", [])
            if not candidates:
                return None
            parts = candidates[0].get("content", {}).get("parts", [])
            return "".join(p.get("text", "") for p in parts).strip()
        except Exception as e:
            log.error("Gemini generateContent call failed: %s", e)
            return None

    def _generate_groq(self, prompt: str, json_mode: bool = True):
        if not self.groq_api_key or requests is None:
            return None
        payload = {
            "model": self.groq_model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
        }
        if json_mode:
            payload["response_format"] = {"type": "json_object"}
        try:
            resp = requests.post(
                GROQ_URL,
                headers={
                    "Authorization": f"Bearer {self.groq_api_key}",
                    "Content-Type": "application/json",
                },
                json=payload,
                timeout=60,
            )
            resp.raise_for_status()
            data = resp.json()
            choices = data.get("choices", [])
            if not choices:
                return None
            return (choices[0].get("message", {}).get("content") or "").strip()
        except Exception as e:
            log.error("Groq chat completion call failed: %s", e)
            return None

    def ask(self, query: str, chunks, top_k: int = 6, provider: str = None, history=None):
        """Returns the raw JSON-string answer from whichever provider is
        configured (Cohere, Gemini, or Groq), or None if none is available /
        the call failed (caller decides the fallback message).

        history: optional list of {"question": str, "answer": str} pairs
        from earlier in this conversation, oldest first — lets the model
        handle follow-ups like "rephrase that" without needing report data
        to answer them."""
        # For small/medium report sets, skip top-k retrieval and send every
        # chunk. Retrieval can silently drop rows needed for sums, counts,
        # or comparisons spanning many rows — correctness matters more than
        # trimming tokens here, and this easily fits typical context windows.
        total_chars = sum(len(c["text"]) for c in chunks)
        MAX_FULL_CONTEXT_CHARS = 150000
        if chunks and total_chars <= MAX_FULL_CONTEXT_CHARS:
            relevant = chunks
        else:
            relevant = self.retrieve(query, chunks, top_k=top_k)
        context = "\n\n---\n\n".join(r["text"] for r in relevant) if relevant else "(no reports attached yet)"
        history_section = _build_history_section(history)
        prompt = _SYSTEM_PROMPT_TMPL.format(history_section=history_section, context=context, query=query)

        chosen = (provider or self.ai_provider or "cohere").lower()
        order = [chosen] + [p for p in ("cohere", "gemini", "groq") if p != chosen]

        for p in order:
            if p == "cohere" and self.co_client:
                result = self._generate_cohere(prompt)
                if result:
                    return _scrub_json_answer(result)
            elif p == "gemini" and self.gemini_api_key:
                result = self._generate_gemini(prompt)
                if result:
                    return _scrub_json_answer(result)
            elif p == "groq" and self.groq_api_key:
                result = self._generate_groq(prompt)
                if result:
                    return _scrub_json_answer(result)
        return None

    def generate_report_html(self, instructions: str, report_context: str, provider: str = None):
        """Sends the raw, real report text through the configured LLM to
        produce a genuinely AI-enhanced dashboard-style HTML report (not a
        mechanical table dump). Returns the raw HTML string, or None if no
        provider is configured / every call failed."""
        instructions = (instructions or "Enhance this report for a professional, executive audience.").strip()
        report_context = (report_context or "").strip() or "(no report data available)"
        prompt = REPORT_ENHANCEMENT_PROMPT_TMPL.format(
            instructions=instructions, report_context=report_context
        )

        chosen = (provider or self.ai_provider or "cohere").lower()
        order = [chosen] + [p for p in ("cohere", "gemini", "groq") if p != chosen]

        for p in order:
            if p == "cohere" and self.co_client:
                result = self._generate_cohere(prompt)
            elif p == "gemini" and self.gemini_api_key:
                result = self._generate_gemini(prompt, timeout=60)
            elif p == "groq" and self.groq_api_key:
                result = self._generate_groq(prompt, json_mode=False)
            else:
                continue
            html = _strip_html_fences(result)
            if html and "<html" in html.lower():
                return html
        return None